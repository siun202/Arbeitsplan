from openpyxl import load_workbook
from datetime import datetime
import sys

# Path to the Excel file
xlsx_path = r"c:\Users\zento\Documents\Titus Köln - Arbeitsplanung - 2026.xlsx"

wb = load_workbook(filename=xlsx_path, data_only=True)
ws = wb.active

# Read header
headers = [str(cell.value).strip() if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]

expected = ['Datum', 'Tag', 'Jerry', 'Caspar', 'Siun', 'Alex', 'Fabian', 'Dominik', 'Chiara', 'Julius']

# Map header indices
header_map = {}
for i, h in enumerate(headers):
    for e in expected:
        if h.lower().startswith(e.lower()[:3]):
            header_map[e] = i
            break

# Fallback: assume columns are in expected order if mapping not complete
if len(header_map) < len(expected):
    header_map = {e: idx for idx, e in enumerate(expected)}

rows_html = []

for row in ws.iter_rows(min_row=2):
    # Read values according to expected order
    cells = [row[header_map[e]].value if header_map.get(e) is not None else None for e in expected]
    # cells[0] should be date
    date_cell = cells[0]
    day_cell = cells[1]

    if date_cell is None and all(c is None for c in cells[2:]):
        continue

    # Format date
    if isinstance(date_cell, datetime):
        date_str = date_cell.strftime('%d.%m.%Y')
    else:
        date_str = str(date_cell) if date_cell is not None else ''

    day_str = str(day_cell).strip() if day_cell is not None else ''

    # Determine if weekend
    tr_class = ' class="weekend"' if day_str.upper() in ('SA', 'SO') else ''

    tds = []
    tds.append(f'    <td class="date-cell">{date_str}</td>')
    tds.append(f'    <td class="day-cell">{day_str}</td>')

    def cell_html(val):
        if val is None or (isinstance(val, str) and val.strip() == ''):
            return '    <td class="free"></td>'
        s = str(val).strip()
        if s == '/':
            return '    <td class="free">/</td>'
        if s.upper() == 'U':
            return '    <td class="vacation">U</td>'
        if s.upper() == 'AT' or s.lower().startswith('feiertag'):
            return '    <td class="special">' + s + '</td>'
        # if looks like time range
        if '-' in s and any(ch.isdigit() for ch in s):
            return '    <td class="work-time">' + s + '</td>'
        # default
        return '    <td class="free">' + s + '</td>'

    for val in cells[2:]:
        tds.append(cell_html(val))

    tr = f'  <tr{tr_class}>\n' + '\n'.join(tds) + '\n  </tr>'
    rows_html.append(tr)

# Write to generated file
out_path = r"c:\Users\zento\Documents\GitHub\Arbeitsplan\generated_rows.html"
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(rows_html))

print(f'Wrote {len(rows_html)} rows to {out_path}')
